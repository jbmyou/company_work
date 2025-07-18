from datetime import datetime
from datetime import timedelta
import calendar
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
import os
from os.path import join
import re
import pandas as pd
from IPython.display import display
    

path_dataset = r"D:\3.자산\전산 dataset"


dtype = {'채무자키': str, '타채무자키': str, '담당자키': str, '관리자비고': str, '계좌키': str, '보증인키': str, '분납키': str, '사건키': str, '신고계좌': str, '입금계좌': str,
 '신용회복키': str, '계좌번호': str, '심의차수': str, '변제금수취계좌': str, '법조치키': str, '관련법조치키': str, '법취하키': str, '타법조치키': str, '관할법원코드': str, '입금키': str, '입금고정키': str, '감면키': str}

    
debt_dtype = {'채무자키':str, '타채무자키':str, '담당자키':str, '관리자비고':str}
account_dtype = {'채무자키':str, '계좌키':str, '타채무자키':str}
grt_dtype = {'채무자키':str, '계좌키':str, '타채무자키':str, '보증인키':str}
rehabilitation_dtype = {'채무자키':str, '계좌키':str, '분납키':str, '사건키':str, '신고계좌':str, '입금계좌':str}
credit_dtype = {"채무자키":str, "계좌키":str, '보증인키':str, '신용회복키' : str, "계좌번호":str, "심의차수":str, "변제금수취계좌":str}
nauri_credit_dtype = {"채무자키":str, "계좌키":str, '보증인키':str, '신용회복키' : str, "계좌번호":str, "심의차수":str, "변제금수취계좌":str}
# credit_dtype = {'채무자키':str, '계좌키':str, '보증인키':str, '심의차수':str, '변제금\n수취계좌':str}
# credit_dtype = {'채무자키' : str, '계좌키' : str, '보증인키':str, '신용회복키' : str, '심의차수' : str, '조정전이율' : str, '조정후이율' : str, '총상환기간' : str, '유예기간' : str, '원금균등상환기간' : str, '원리균등상환기간' : str, '이자상환기간' : str, '납입회차' : str, '연체기간' : str, '원금균등시작회차' : str, '원금균등종료회차' : str, '원리균등시작회차' : str, '원리균등종료회차' : str, '이자상환시작회차' : str, '이자상환종료회차' : str, '감면율' : str, '주채무자주민번호' : str, '시작회차1' : str, '종료회차1' : str, '채무조정1적용이율1' : str, '채무조정2적용이율1' : str, '시작회차2' : str, '종료회차2' : str, '채무조정1적용이율2' : str, '채무조정2적용이율2' : str, '시작회차3' : str, '종료회차3' : str, '채무조정1적용이율3' : str, '채무조정2적용이율3' : str, '시작회차4' : str, '종료회차4' : str, '채무조정1적용이율4' : str, '채무조정2적용이율4' : str, '시작회차5' : str, '종료회차5' : str, '채무조정1적용이율5' : str, '채무조정2적용이율5' : str, '시작회차6' : str, '종료회차6' : str, '채무조정1적용이율6' : str, '채무조정2적용이율6' : str, '시작회차7' : str, '종료회차7' : str, '채무조정1적용이율7' : str, '채무조정2적용이율7' : str, '시작회차8' : str, '종료회차8' : str, '채무조정1적용이율8' : str, '채무조정2적용이율8' : str, '시작회차9' : str, '종료회차9' : str, '채무조정1적용이율9' : str, '채무조정2적용이율9' : str, '시작회차10' : str, '종료회차10' : str, '채무조정1적용이율10' : str, '채무조정2적용이율10' : str, '재조정횟수' : str, '수정조정횟수' : str, '일시납감면율' : str, '거치기간' : str, '체증전구간의시작회차' : str, '체증전구간의종료회차' : str, '체증전구간의기준기간' : str, '변제금수취계좌' : str, '담보권실행유예기간':str
# }
event_dtype = {'채무자키':str, '법조치키':str, '계좌키':str, '관련법조치키':str, '법취하키':str, '타법조치키':str, '타채무자키':str, '관할법원코드':str}
deposit_dtype = {'채무자키':str, '입금키':str, '계좌키':str, '계좌번호':str, '입금고정키':str, '타채무자키':str}
reduction_dtype = {'채무자키':str, '계좌키':str, '감면키':str}
installment_dtype = {'채무자키':str, '계좌키':str, '분납키':str}


def save_df_to_excel_underline(df, fullpath, key_columns_no=1, font_size=9):
    """키값의 열번호나 열이름을 작성하면 키값이 달라질때마다 밑줄 그어줌(a열이 1번), 필요없으면 0입력"""
    
    # na값 처리 : onpenpyxl은 pandas.NA값을 처리 못함
    df = df.replace({pd.NA:None})
    
    
    # 엑셀 열기
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 시트에 데이터 쓰기
    for r in dataframe_to_rows(df, index=False, header=True) : 
        ws.append(r)
        
    # 스타일 지정
    # 폰트 (데이터 부분)
    font_data = Font(size=font_size)
    # 배경색 (헤더)
    fill_col = PatternFill(fill_type='solid', start_color='FFDDD9C4', end_color='FFDDD9C4') # 칼럼명
    alignment = Alignment(horizontal='center', vertical='center')
    
    # 헤더행 스타일 적용    
    for cell in ws[1] :
        cell.font = Font(size=font_size, bold=True)
        cell.fill = fill_col
        cell.alignment = alignment
        
    # 데이터행 스타일 적용
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            # 정수 데이터인 경우
            if isinstance(cell.value, int):
                cell.font = font_data
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            # 날짜 데이터인 경우
            elif isinstance(cell.value, pd.Timestamp):
                cell.font = font_data
                cell.number_format = 'yyyy-mm-dd'
            # 나머지 문자열 데이터인 경우
            else:
                cell.font = font_data
     
    
    # 그룹 열의 값이 달라질 때마다 밑줄 추가
        # 칼럼명이 주어진 경우 칼럼순번으로 바꾸기
    if isinstance(key_columns_no, str) :
        key_columns_no = df.columns.get_loc(key_columns_no) + 1
    
    if key_columns_no > 0 : 
        underline_border = Border(bottom=Side(style='thin'))
        prev_value = None
                
        for row in range(2, ws.max_row + 1):
            current_value = ws.cell(row=row, column=key_columns_no).value  # 주민등록번호가 첫 번째 열에 있다고 가정
            if prev_value is not None and current_value != prev_value:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row-1, column=col).border = underline_border
            prev_value = current_value        
    

    # 엑셀 파일 저장
    wb.save(fullpath)
    
def save_df_to_excel_two_underline(df, fullpath, key_columns_no=1, sub_key_columns_no=0, font_size=9):
    """키값의 열번호나 열이름을 작성하면 키값이 달라질때마다 밑줄 그어줌(a열이 1번), 필요없으면 0입력"""
    
    # na값 처리 : onpenpyxl은 pandas.NA값을 처리 못함
    df = df.replace({pd.NA:None})
    
    
    # 엑셀 열기
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 시트에 데이터 쓰기
    for r in dataframe_to_rows(df, index=False, header=True) : 
        ws.append(r)
        
    # 스타일 지정
    # 폰트 (데이터 부분)
    font_data = Font(size=font_size)
    # 배경색 (헤더)
    fill_col = PatternFill(fill_type='solid', start_color='FFDDD9C4', end_color='FFDDD9C4') # 칼럼명
    alignment = Alignment(horizontal='center', vertical='center')
    
    # 헤더행 스타일 적용    
    for cell in ws[1] :
        cell.font = Font(size=font_size, bold=True)
        cell.fill = fill_col
        cell.alignment = alignment

    # 데이터행 스타일 적용
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            # 정수 데이터인 경우
            if isinstance(cell.value, int):
                cell.font = font_data
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            # 날짜 데이터인 경우
            elif isinstance(cell.value, pd.Timestamp):
                cell.font = font_data
                cell.number_format = 'yyyy-mm-dd'
            # 나머지 문자열 데이터인 경우
            else:
                cell.font = font_data
     
    
    # 밑줄 스타일 지정
    underline_border = Border(bottom=Side(style='thin'))
    dotted_border = Border(bottom=Side(style='dotted'))
    
    prev_value = None
    prev_sub_value = None
    
    
    # 그룹 열의 값이 달라질 때마다 밑줄 추가
        # 칼럼명이 주어진 경우 칼럼순번으로 바꾸기
    if isinstance(key_columns_no, str) :
        key_columns_no = df.columns.get_loc(key_columns_no) + 1
        
    if isinstance(sub_key_columns_no, str) :
        sub_key_columns_no = df.columns.get_loc(sub_key_columns_no) + 1
        
    for row in range(2, ws.max_row + 1):
        current_value = ws.cell(row=row, column=key_columns_no).value if key_columns_no > 0 else None
        sub_value = ws.cell(row=row, column=sub_key_columns_no).value if sub_key_columns_no > 0 else None
        
        if prev_value is not None and current_value != prev_value:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row - 1, column=col).border = underline_border
        elif prev_sub_value is not None and sub_value != prev_sub_value:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row - 1, column=col).border = dotted_border
        
        prev_value = current_value
        prev_sub_value = sub_value 
    

    # 엑셀 파일 저장
    wb.save(fullpath)


def save_excel_with_explain(df, fullpath, key_columns_no=1, sub_key_columns_no=0, font_size=9) :
    save_fn = os.path.basename(fullpath).split('.')[0]
    if len(df)>0 :
        print(f'★ {save_fn} {len(df)}건 엑셀출력')
        save_df_to_excel_two_underline(df, fullpath, key_columns_no=key_columns_no, sub_key_columns_no=sub_key_columns_no, font_size=font_size)
        print('')
    else : 
        print(f'☆ {save_fn} 출력건 없음')
        print('')


def display_with_explain(df, explain="") : 
    if not df.empty :
        print(f'★ {explain} {len(df)}건')
        display(df)
        print('')
    else : 
        print(f'☆ {explain} 없음')
        print('')


# 차감할연체이자 계산용
def 문자열날짜차이(입금일:str, 기준일:str) :
    date_format = "%Y-%m-%d"
    """ 문자열 타입의 두 날짜 사이의 날짜차이 리턴 """
    date1 = datetime.strptime(입금일, date_format)
    date2 = datetime.strptime(기준일, date_format)
    return (date2 - date1).days # .days : 정수값 얻기위해


def 키워드로파일명찾기(폴더:str, 포함키워드:str, 제외키워드="", 전체경로=True, 여러파일허용=False) :
    """포함키워드, 제외키워드로 파일 경로 찾기, 여러파일 허용시 목록을 반환하므로 선택할 수 있음"""
    file_list = [file for file in os.listdir(폴더) if not file.startswith('~$')]
    if 제외키워드 == "" : 
        fn = [file for file in file_list if re.search(포함키워드, file)]
    else : 
        fn = [file for file in file_list if (re.search(포함키워드, file)!=None) & (re.search(제외키워드,file)==None)]
    
    if len(fn) == 1 :
        if 전체경로 : 
            return join(폴더,fn[0])
        else : 
            return fn[0]
    elif len(fn) == 0 :
        print(f"포함키워드:{포함키워드}, 제외키워드{제외키워드} 조건을 만족하는 파일이 없습니다.")
    else : 
        if 여러파일허용 : 
            if 전체경로 : 
                return [join(폴더,f) for f in fn]
            else : 
                return fn
        else : 
            print(f"포함키워드:{포함키워드}, 제외키워드{제외키워드} 조건을 만족하는 파일이 둘 이상입니다.")
            print(fn)


            
def swap_columns(df, col1, col2):
    cols = df.columns.tolist()
    idx1, idx2 = cols.index(col1), cols.index(col2)
    cols[idx1], cols[idx2] = cols[idx2], cols[idx1]
    return df[cols]
